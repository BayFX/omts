#!/usr/bin/env python3
"""Generate the OMTS Excel import template and example file.

This script produces two files:
  - omts-import-template.xlsx: Empty template with data validation and headers
  - omts-import-example.xlsx: Template populated with the SPEC-001 Section 10 example data

Both files use the multi-sheet structure defined by the Excel Import Format
specification, aligned with the expert panel recommendations.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
METADATA_KEY_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, num_cols):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def add_data_validation(ws, col_letter, values, prompt_title="", prompt_msg=""):
    """Add a dropdown data validation to a column."""
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = prompt_msg
    dv.promptTitle = prompt_title
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid value"
    dv.error = f"Must be one of: {', '.join(values)}"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}10000")


def set_col_widths(ws, widths):
    """Set column widths from a dict of {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def add_header_comments(ws, comments, header_row=1):
    """Attach tooltip comments to header cells."""
    for col_idx, text in comments.items():
        cell = ws.cell(row=header_row, column=col_idx)
        cell.comment = Comment(text, "OMTS")


# ── Sheet definitions ────────────────────────────────────────────────────────

def create_metadata_sheet(wb):
    """Create the Metadata sheet with file-level fields."""
    ws = wb.active
    ws.title = "Metadata"

    fields = [
        ("Field", "Value", "Description"),
        ("snapshot_date", "", "ISO 8601 date (YYYY-MM-DD) when this snapshot was produced. REQUIRED."),
        ("reporting_entity", "", "ID of the organization node whose perspective this file represents (optional)."),
        ("disclosure_scope", "", "Intended audience: internal, partner, or public (optional)."),
        ("default_confidence", "", "Default data quality confidence: verified, reported, inferred, estimated (optional)."),
        ("default_source", "", "Default data quality source description (optional)."),
        ("default_last_verified", "", "Default date data was last verified, ISO 8601 (optional)."),
    ]

    for row_idx, (field, value, desc) in enumerate(fields, start=1):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=desc)

        if row_idx == 1:
            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
        else:
            ws.cell(row=row_idx, column=1).font = METADATA_KEY_FONT
            ws.cell(row=row_idx, column=3).alignment = WRAP_ALIGNMENT
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Data validation for disclosure_scope
    dv = DataValidation(type="list", formula1='"internal,partner,public"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B4")

    # Data validation for default_confidence
    dv2 = DataValidation(
        type="list", formula1='"verified,reported,inferred,estimated"', allow_blank=True
    )
    ws.add_data_validation(dv2)
    dv2.add("B5")

    set_col_widths(ws, {"A": 22, "B": 30, "C": 70})

    metadata_comments = {
        2: "Required. ISO 8601 date (YYYY-MM-DD) when this data snapshot was produced.",
        3: "Node ID of the organization whose perspective this file represents. Must match an id in the Organizations sheet.",
        4: "Who will see this file: internal (company only), partner (supply-chain partners), or public.",
        5: "Default data quality level: verified, reported, inferred, or estimated.",
        6: 'Default description of how data was collected (e.g. "manual-review", "erp-export").',
        7: "ISO 8601 date when data was last verified.",
    }
    for row, text in metadata_comments.items():
        ws.cell(row=row, column=1).comment = Comment(text, "OMTS")

    return ws


def create_organizations_sheet(wb):
    """Create the Organizations sheet."""
    ws = wb.create_sheet("Organizations")

    headers = [
        "id",                       # A - graph-local ID (auto-generated if blank)
        "name",                     # B - REQUIRED
        "jurisdiction",             # C - ISO 3166-1 alpha-2
        "status",                   # D - active/dissolved/merged/suspended
        "lei",                      # E - LEI (20-char)
        "duns",                     # F - DUNS (9-digit)
        "nat_reg_value",            # G - national registry number
        "nat_reg_authority",        # H - GLEIF RA code
        "vat_value",                # I - VAT/tax ID
        "vat_country",              # J - ISO 3166-1 alpha-2
        "internal_id",              # K - internal system ID
        "internal_system",          # L - internal system name (authority)
        "risk_tier",                # M - label: risk-tier
        "kraljic_quadrant",         # N - label: kraljic-quadrant
        "approval_status",          # O - label: approval-status
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    # Data validations
    add_data_validation(ws, "D", ["active", "dissolved", "merged", "suspended"],
                        "Status", "Organization lifecycle status")
    add_data_validation(ws, "M", ["critical", "high", "medium", "low"],
                        "Risk Tier", "General risk classification")
    add_data_validation(ws, "N", ["strategic", "leverage", "bottleneck", "non-critical"],
                        "Kraljic Quadrant", "Kraljic portfolio classification")
    add_data_validation(ws, "O", ["approved", "conditional", "pending", "blocked", "phase-out"],
                        "Approval Status", "Supplier approval status")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 14, "D": 12, "E": 24, "F": 14,
        "G": 20, "H": 18, "I": 20, "J": 12, "K": 16, "L": 20,
        "M": 12, "N": 16, "O": 16,
    })

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated from name if left blank.",
        2: "Required. Legal name of the organization.",
        3: "ISO 3166-1 alpha-2 country code of incorporation.",
        4: "Organization lifecycle: active, dissolved, merged, or suspended.",
        5: "Legal Entity Identifier (20 characters, ISO 17442).",
        6: "D-U-N-S Number (9 digits).",
        7: "National business registry number (e.g. Companies House number).",
        8: "GLEIF Registration Authority code (e.g. RA000585 for UK Companies House).",
        9: "VAT or tax identification number.",
        10: "ISO 3166-1 alpha-2 country that issued the VAT number.",
        11: "Your internal system identifier (e.g. SAP vendor number).",
        12: 'Name of the internal system (e.g. "sap-mm-prod").',
        13: "General risk classification: critical, high, medium, or low.",
        14: "Kraljic portfolio classification: strategic, leverage, bottleneck, or non-critical.",
        15: "Supplier approval status: approved, conditional, pending, blocked, or phase-out.",
    })

    return ws


def create_facilities_sheet(wb):
    """Create the Facilities sheet."""
    ws = wb.create_sheet("Facilities")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "operator_id",              # C - ref to org id
        "address",                  # D
        "latitude",                 # E
        "longitude",                # F
        "gln",                      # G - GLN (13-digit)
        "internal_id",              # H
        "internal_system",          # I
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {
        "A": 20, "B": 30, "C": 18, "D": 40, "E": 14, "F": 14,
        "G": 18, "H": 16, "I": 20,
    })

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated from name if blank.",
        2: 'Required. Name of the facility (e.g. "Sheffield Plant").',
        3: "ID of the organization that operates this facility. Must match an id in Organizations.",
        4: "Street address or location description.",
        5: "WGS 84 latitude (decimal degrees, e.g. 53.3811).",
        6: "WGS 84 longitude (decimal degrees, e.g. -1.4701).",
        7: "Global Location Number (13 digits, GS1).",
        8: "Your internal site identifier.",
        9: "Name of the internal system.",
    })

    return ws


def create_goods_sheet(wb):
    """Create the Goods sheet."""
    ws = wb.create_sheet("Goods")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "commodity_code",           # C - HS/CN code
        "unit",                     # D - e.g., kg, mt, pcs
        "gtin",                     # E - GTIN
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {"A": 20, "B": 30, "C": 16, "D": 10, "E": 18})

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated from name if blank.",
        2: "Required. Name of the product or material.",
        3: "HS or CN commodity code (e.g. 7318.15 for bolts).",
        4: "Unit of measure (e.g. kg, mt, pcs).",
        5: "Global Trade Item Number (GS1).",
    })

    return ws


def create_attestations_sheet(wb):
    """Create the Attestations sheet."""
    ws = wb.create_sheet("Attestations")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "attestation_type",         # C - REQUIRED
        "standard",                 # D
        "issuer",                   # E
        "valid_from",               # F - REQUIRED (YYYY-MM-DD)
        "valid_to",                 # G
        "outcome",                  # H
        "status",                   # I
        "reference",                # J
        "risk_severity",            # K
        "risk_likelihood",          # L
        "attested_entity_id",       # M - which node this attests (for attested_by edge)
        "scope",                    # N - attested_by edge scope
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "C",
                        ["certification", "audit", "due_diligence_statement", "self_declaration", "other"],
                        "Attestation Type", "Type of attestation")
    add_data_validation(ws, "H",
                        ["pass", "conditional_pass", "fail", "pending", "not_applicable"],
                        "Outcome", "Attestation outcome")
    add_data_validation(ws, "I",
                        ["active", "suspended", "revoked", "expired", "withdrawn"],
                        "Status", "Attestation lifecycle status")
    add_data_validation(ws, "K",
                        ["critical", "high", "medium", "low"],
                        "Risk Severity", "Risk severity classification")
    add_data_validation(ws, "L",
                        ["very_likely", "likely", "possible", "unlikely"],
                        "Risk Likelihood", "Risk likelihood")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 24, "D": 20, "E": 30, "F": 14,
        "G": 14, "H": 16, "I": 12, "J": 20, "K": 14, "L": 14,
        "M": 22, "N": 20,
    })

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated if blank.",
        2: "Required. Name of the certification or audit.",
        3: "Required. Type: certification, audit, due_diligence_statement, self_declaration, or other.",
        4: "Standard or scheme (e.g. SA8000:2014, ISO 14001).",
        5: "Name of the issuing body.",
        6: "Required. ISO 8601 date (YYYY-MM-DD) when the attestation takes effect.",
        7: "ISO 8601 date when the attestation expires.",
        8: "Result: pass, conditional_pass, fail, pending, or not_applicable.",
        9: "Lifecycle: active, suspended, revoked, expired, or withdrawn.",
        10: "External reference number or URL.",
        11: "Risk severity: critical, high, medium, or low.",
        12: "Risk likelihood: very_likely, likely, possible, or unlikely.",
        13: "Node ID of the entity this attestation covers. Creates an attested_by edge.",
        14: "Scope of the attestation (e.g. \"working conditions\", \"environmental compliance\").",
    })

    return ws


def create_consignments_sheet(wb):
    """Create the Consignments sheet."""
    ws = wb.create_sheet("Consignments")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "lot_id",                   # C
        "quantity",                 # D
        "unit",                     # E
        "production_date",          # F
        "origin_country",           # G - ISO 3166-1 alpha-2
        "installation_id",          # H - ref to facility
        "direct_emissions_co2e",    # I - CBAM
        "indirect_emissions_co2e",  # J - CBAM
        "emission_factor_source",   # K - CBAM
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "K",
                        ["actual", "default_eu", "default_country"],
                        "Emission Factor", "Source of emissions data")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 14, "D": 12, "E": 10, "F": 16,
        "G": 16, "H": 18, "I": 22, "J": 24, "K": 24,
    })

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated if blank.",
        2: "Required. Description of the consignment or batch.",
        3: "Lot or batch identifier.",
        4: "Quantity in the consignment.",
        5: "Unit of measure (e.g. kg, mt).",
        6: "ISO 8601 date of production.",
        7: "ISO 3166-1 alpha-2 country of origin.",
        8: "ID of the producing facility. Must match an id in Facilities.",
        9: "Direct (Scope 1) emissions in tonnes CO2e (CBAM).",
        10: "Indirect (Scope 2) emissions in tonnes CO2e (CBAM).",
        11: "Source of emissions data: actual, default_eu, or default_country.",
    })

    return ws


def create_supply_relationships_sheet(wb):
    """Create the Supply Relationships sheet for supply-chain edges."""
    ws = wb.create_sheet("Supply Relationships")

    headers = [
        "id",                       # A - edge ID (auto-generated if blank)
        "type",                     # B - supplies/subcontracts/tolls/distributes/brokers/sells_to/operates/produces
        "supplier_id",              # C - source node (supplier/operator/facility)
        "buyer_id",                 # D - target node (buyer/facility/good)
        "valid_from",               # E - REQUIRED (YYYY-MM-DD)
        "valid_to",                 # F
        "commodity",                # G - HS code or description
        "tier",                     # H - tier relative to reporting_entity
        "volume",                   # I
        "volume_unit",              # J
        "annual_value",             # K
        "value_currency",           # L - ISO 4217
        "contract_ref",             # M
        "share_of_buyer_demand",    # N - 0-100
        "service_type",             # O - for distributes edges
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["supplies", "subcontracts", "tolls", "distributes", "brokers",
                         "sells_to", "operates", "produces", "composed_of"],
                        "Edge Type", "Type of supply/operational relationship")
    add_data_validation(ws, "O",
                        ["warehousing", "transport", "fulfillment", "other"],
                        "Service Type", "For distributes edges only")

    set_col_widths(ws, {
        "A": 14, "B": 16, "C": 18, "D": 18, "E": 14, "F": 14,
        "G": 16, "H": 8, "I": 12, "J": 14, "K": 14, "L": 14,
        "M": 16, "N": 22, "O": 16,
    })

    add_header_comments(ws, {
        1: "Edge identifier. Auto-generated if blank.",
        2: "Required. Relationship type: supplies, subcontracts, tolls, distributes, brokers, sells_to, operates, produces, composed_of.",
        3: "Required. Source node ID (the supplier, operator, or facility).",
        4: "Required. Target node ID (the buyer, facility, or good).",
        5: "Required. ISO 8601 date when the relationship started.",
        6: "ISO 8601 date when the relationship ended.",
        7: "HS/CN code or description of what is supplied.",
        8: "Supply-chain tier relative to the reporting entity (1 = direct).",
        9: "Quantity supplied.",
        10: "Unit for volume (e.g. kg, mt, pcs).",
        11: "Annual monetary value of the relationship.",
        12: "ISO 4217 currency code.",
        13: "Contract or agreement reference number.",
        14: "Percentage of buyer's demand met by this supplier (0-100).",
        15: "For distributes edges only: warehousing, transport, fulfillment, or other.",
    })

    return ws


def create_corporate_structure_sheet(wb):
    """Create the Corporate Structure sheet for hierarchy edges."""
    ws = wb.create_sheet("Corporate Structure")

    headers = [
        "id",                       # A
        "type",                     # B - ownership/legal_parentage/operational_control/beneficial_ownership
        "subsidiary_id",            # C - source (child/subsidiary/person)
        "parent_id",                # D - target (parent/organization)
        "valid_from",               # E - REQUIRED
        "valid_to",                 # F
        "percentage",               # G - for ownership/beneficial_ownership
        "direct",                   # H - TRUE/FALSE
        "control_type",             # I - for operational_control/beneficial_ownership
        "consolidation_basis",      # J - for legal_parentage
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["ownership", "legal_parentage", "operational_control", "beneficial_ownership"],
                        "Edge Type", "Type of corporate relationship")
    add_data_validation(ws, "H", ["TRUE", "FALSE"],
                        "Direct", "Direct or indirect relationship")
    add_data_validation(ws, "I",
                        ["franchise", "management", "tolling", "licensed_manufacturing", "other",
                         "voting_rights", "capital", "other_means", "senior_management"],
                        "Control Type", "For operational_control or beneficial_ownership")
    add_data_validation(ws, "J",
                        ["ifrs10", "us_gaap_asc810", "other", "unknown"],
                        "Consolidation Basis", "For legal_parentage only")

    set_col_widths(ws, {
        "A": 14, "B": 24, "C": 18, "D": 18, "E": 14, "F": 14,
        "G": 14, "H": 10, "I": 26, "J": 22,
    })

    add_header_comments(ws, {
        1: "Edge identifier. Auto-generated if blank.",
        2: "Required. Relationship: ownership, legal_parentage, operational_control, or beneficial_ownership.",
        3: "Required. The child/subsidiary entity node ID.",
        4: "Required. The parent entity node ID.",
        5: "Required. ISO 8601 date when the relationship started.",
        6: "ISO 8601 date when the relationship ended.",
        7: "Ownership or control percentage (0-100). For ownership and beneficial_ownership.",
        8: "TRUE for direct relationships, FALSE for indirect.",
        9: "Type of control (for operational_control or beneficial_ownership).",
        10: "Accounting consolidation basis (for legal_parentage): ifrs10, us_gaap_asc810, other, unknown.",
    })

    return ws


def create_persons_sheet(wb):
    """Create the Persons sheet for beneficial owners and key individuals."""
    ws = wb.create_sheet("Persons")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "jurisdiction",             # C - ISO 3166-1 alpha-2
        "role",                     # D
        "nationality",              # E - ISO 3166-1 alpha-2
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {"A": 18, "B": 30, "C": 14, "D": 20, "E": 14})

    add_header_comments(ws, {
        1: "Graph-local identifier. Auto-generated if blank.",
        2: "Required. Full name of the individual.",
        3: "ISO 3166-1 alpha-2 country of residence.",
        4: 'Role description (e.g. "Director", "UBO").',
        5: "ISO 3166-1 alpha-2 nationality.",
    })

    return ws


def create_same_as_sheet(wb):
    """Create the Same As sheet for entity deduplication assertions."""
    ws = wb.create_sheet("Same As")

    headers = [
        "entity_a",                 # A - ref to node ID
        "entity_b",                 # B - ref to node ID
        "confidence",               # C - definite/probable/possible
        "basis",                    # D - justification (e.g., name_match, manual_review)
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "C",
                        ["definite", "probable", "possible"],
                        "Confidence", "Confidence level of the equivalence assertion")

    set_col_widths(ws, {"A": 20, "B": 20, "C": 14, "D": 40})

    add_header_comments(ws, {
        1: "Node ID of the first entity.",
        2: "Node ID of the second entity (asserted to be the same real-world entity).",
        3: "Confidence level: definite, probable, or possible.",
        4: 'Justification for the assertion (e.g. "LEI match", "manual review").',
    })

    return ws


def create_identifiers_sheet(wb):
    """Create the Identifiers sheet for advanced multi-identifier scenarios."""
    ws = wb.create_sheet("Identifiers")

    headers = [
        "node_id",                  # A - ref to node in any sheet
        "scheme",                   # B - lei/duns/gln/nat-reg/vat/internal/extension
        "value",                    # C
        "authority",                # D - required for nat-reg, vat, internal
        "sensitivity",              # E - public/restricted/confidential
        "valid_from",               # F
        "valid_to",                 # G
        "verification_status",      # H
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["lei", "duns", "gln", "nat-reg", "vat", "internal"],
                        "Scheme", "Identifier scheme")
    add_data_validation(ws, "E",
                        ["public", "restricted", "confidential"],
                        "Sensitivity", "Identifier sensitivity level")
    add_data_validation(ws, "H",
                        ["verified", "reported", "inferred", "unverified"],
                        "Verification", "Verification status")

    set_col_widths(ws, {
        "A": 18, "B": 12, "C": 24, "D": 20, "E": 14, "F": 14,
        "G": 14, "H": 18,
    })

    add_header_comments(ws, {
        1: "ID of the node this identifier belongs to. Must match an id in any entity sheet.",
        2: "Identifier scheme: lei, duns, gln, nat-reg, vat, or internal.",
        3: "The identifier value.",
        4: "Issuing authority. Required for nat-reg, vat, and internal schemes.",
        5: "Access level: public, restricted, or confidential.",
        6: "ISO 8601 date when the identifier became valid.",
        7: "ISO 8601 date when the identifier expired.",
        8: "Verification state: verified, reported, inferred, or unverified.",
    })

    return ws


def create_readme_sheet(wb):
    """Create a README sheet with instructions."""
    ws = wb.create_sheet("README")

    instructions = [
        ("OMTS Excel Import Template", ""),
        ("", ""),
        ("This workbook is designed for import into the OMTS format using:", ""),
        ("    omtsf import-excel <this-file.xlsx> -o output.omts", ""),
        ("", ""),
        ("SHEET OVERVIEW", ""),
        ("Metadata", "File-level settings: snapshot date, reporting entity, disclosure scope."),
        ("Organizations", "Legal entities (companies, NGOs, government bodies)."),
        ("Facilities", "Physical locations (factories, warehouses, farms, mines)."),
        ("Goods", "Products, materials, or commodities."),
        ("Persons", "Beneficial owners, key individuals (sensitivity: confidential by default)."),
        ("Attestations", "Certifications, audits, due diligence statements."),
        ("Consignments", "Batches, lots, shipments (optional, for CBAM/EUDR)."),
        ("Supply Relationships", "Supply, subcontracting, tolling, distribution edges."),
        ("Corporate Structure", "Ownership, legal parentage, operational control edges."),
        ("Same As", "Entity deduplication: link nodes that represent the same real-world entity."),
        ("Identifiers", "Advanced: additional identifiers beyond the common columns."),
        ("", ""),
        ("REQUIRED FIELDS", ""),
        ("Organizations", "name"),
        ("Facilities", "name"),
        ("Goods", "name"),
        ("Attestations", "name, attestation_type, valid_from"),
        ("Supply Relationships", "type, supplier_id, buyer_id, valid_from"),
        ("Corporate Structure", "type, subsidiary_id, parent_id, valid_from"),
        ("", ""),
        ("AUTO-GENERATED FIELDS", ""),
        ("The import command will auto-generate:", ""),
        ("  - file_salt (cryptographic random)", ""),
        ("  - node/edge IDs (if left blank)", ""),
        ("  - boundary_ref nodes (if disclosure_scope set)", ""),
        ("  - sensitivity defaults per SPEC-004", ""),
        ("", ""),
        ("IDENTIFIER COLUMNS", ""),
        ("Common identifiers have dedicated columns on the Organizations sheet:", ""),
        ("  lei          - Legal Entity Identifier (20-char, validated)", ""),
        ("  duns         - DUNS Number (9-digit, validated)", ""),
        ("  nat_reg_*    - National registry number + GLEIF RA authority code", ""),
        ("  vat_*        - VAT/tax ID + ISO 3166-1 alpha-2 country code", ""),
        ("  internal_*   - Internal system ID + system name", ""),
        ("For multiple IDs of the same scheme, use the Identifiers sheet.", ""),
        ("", ""),
        ("EDGE DIRECTION", ""),
        ("Supply Relationships: supplier_id = who supplies, buyer_id = who buys", ""),
        ("Corporate Structure: subsidiary_id = child entity, parent_id = parent entity", ""),
        ("", ""),
        ("ENTITY DEDUPLICATION", ""),
        ("Use the Same As sheet to link nodes that represent the same real-world entity", ""),
        ("but appear as separate rows (e.g., same company under different names/IDs).", ""),
        ("The import command uses these to generate same_as edges for merge operations.", ""),
        ("", ""),
        ("PERSON NODE PRIVACY", ""),
        ("Person nodes default to confidential sensitivity (SPEC-004).", ""),
        ("If disclosure_scope is 'public', the import command will reject the file", ""),
        ("if any person nodes are present.", ""),
        ("", ""),
        ("SPEC VERSION", ""),
        ("This template targets OMTS spec version 0.1.0", ""),
    ]

    title_font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    normal_font = Font(name="Calibri", size=11)
    code_font = Font(name="Consolas", size=10)

    for row_idx, (col_a, col_b) in enumerate(instructions, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws.cell(row=row_idx, column=2, value=col_b)

        if row_idx == 1:
            cell_a.font = title_font
        elif col_a.isupper() and col_a.strip():
            cell_a.font = section_font
        elif col_a.startswith("    ") or col_a.startswith("  -"):
            cell_a.font = code_font
        else:
            cell_a.font = normal_font

        cell_b.font = normal_font

    set_col_widths(ws, {"A": 50, "B": 60})

    # Move README to first position
    wb.move_sheet("README", offset=-wb.sheetnames.index("README"))
    return ws


# ── Example data ─────────────────────────────────────────────────────────────

def populate_example_data(wb):
    """Populate with SPEC-001 Section 10 example data (Acme-Bolt scenario)."""

    # Metadata
    ws = wb["Metadata"]
    ws["B2"] = "2026-02-17"
    ws["B3"] = "org-acme"
    ws["B4"] = "partner"
    ws["B5"] = "reported"
    ws["B6"] = "manual-review"
    ws["B7"] = "2026-02-17"

    # Organizations
    ws = wb["Organizations"]
    # Row 2: Acme Manufacturing GmbH
    ws["A2"] = "org-acme"
    ws["B2"] = "Acme Manufacturing GmbH"
    ws["C2"] = "DE"
    ws["D2"] = "active"
    ws["E2"] = "5493006MHB84DD0ZWV18"
    ws["F2"] = "081466849"
    ws["G2"] = "HRB86891"
    ws["H2"] = "RA000548"
    ws["I2"] = "DE123456789"
    ws["J2"] = "DE"
    ws["K2"] = "V-100234"
    ws["L2"] = "sap-mm-prod"

    # Row 3: Bolt Supplies Ltd
    ws["A3"] = "org-bolt"
    ws["B3"] = "Bolt Supplies Ltd"
    ws["C3"] = "GB"
    ws["D3"] = "active"
    ws["F3"] = "234567890"
    ws["G3"] = "07228507"
    ws["H3"] = "RA000585"
    ws["M3"] = "low"
    ws["N3"] = "strategic"

    # Facilities
    ws = wb["Facilities"]
    ws["A2"] = "fac-bolt-sheffield"
    ws["B2"] = "Bolt Sheffield Plant"
    ws["C2"] = "org-bolt"
    ws["E2"] = 53.3811
    ws["F2"] = -1.4701

    # Goods
    ws = wb["Goods"]
    ws["A2"] = "good-steel-bolts"
    ws["B2"] = "M10 Steel Hex Bolts"
    ws["C2"] = "7318.15"

    # Attestations
    ws = wb["Attestations"]
    ws["A2"] = "att-sa8000"
    ws["B2"] = "SA8000 Certification"
    ws["C2"] = "certification"
    ws["D2"] = "SA8000:2014"
    ws["E2"] = "Social Accountability International"
    ws["F2"] = "2025-06-01"
    ws["G2"] = "2028-05-31"
    ws["H2"] = "pass"
    ws["I2"] = "active"
    ws["M2"] = "fac-bolt-sheffield"
    ws["N2"] = "working conditions"

    # Supply Relationships
    ws = wb["Supply Relationships"]
    # Edge: Bolt supplies Acme
    ws["A2"] = "edge-001"
    ws["B2"] = "supplies"
    ws["C2"] = "org-bolt"
    ws["D2"] = "org-acme"
    ws["E2"] = "2023-01-15"
    ws["G2"] = "7318.15"
    ws["H2"] = 1

    # Edge: Bolt operates Sheffield plant
    ws["A3"] = "edge-002"
    ws["B3"] = "operates"
    ws["C3"] = "org-bolt"
    ws["D3"] = "fac-bolt-sheffield"
    ws["E3"] = "2018-06-01"

    # Edge: Sheffield produces steel bolts
    ws["A4"] = "edge-003"
    ws["B4"] = "produces"
    ws["C4"] = "fac-bolt-sheffield"
    ws["D4"] = "good-steel-bolts"
    ws["E4"] = "2020-03-01"

    # Corporate Structure
    ws = wb["Corporate Structure"]
    # Edge: Acme owns 51% of Bolt
    ws["A2"] = "edge-004"
    ws["B2"] = "ownership"
    ws["C2"] = "org-acme"
    ws["D2"] = "org-bolt"
    ws["E2"] = "2019-04-01"
    ws["G2"] = 51.0

    # Identifiers (additional - GLN for facility, GTIN for good)
    ws = wb["Identifiers"]
    ws["A2"] = "fac-bolt-sheffield"
    ws["B2"] = "gln"
    ws["C2"] = "5060012340001"
    ws["E2"] = "public"

    ws["A3"] = "fac-bolt-sheffield"
    ws["B3"] = "internal"
    ws["C3"] = "SITE-SHF-01"
    ws["D3"] = "bolt-erp"
    ws["E3"] = "restricted"

    ws["A4"] = "good-steel-bolts"
    ws["B4"] = "org.gs1.gtin"
    ws["C4"] = "05060012340018"
    ws["E4"] = "public"


# ── Supplier List (simplified single-sheet template) ────────────────────────

SUPPLIER_LIST_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
METADATA_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="2F5496")
METADATA_VALUE_FONT = Font(name="Calibri", size=10)
METADATA_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def create_supplier_list_workbook():
    """Create the simplified single-sheet OMTS supplier list workbook.

    Layout:
      Row 1-2: Metadata key-value pairs (reporting entity, snapshot date)
      Row 3:   Blank separator
      Row 4:   Column headers
      Row 5+:  Data rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier List"

    # ── Metadata area (rows 1-2) ────────────────────────────────────────────

    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = METADATA_FILL
        ws.cell(row=2, column=col).fill = METADATA_FILL

    ws.cell(row=1, column=1, value="Reporting Entity").font = METADATA_LABEL_FONT
    ws.cell(row=1, column=2).font = METADATA_VALUE_FONT
    ws.cell(row=1, column=3, value="Snapshot Date").font = METADATA_LABEL_FONT
    ws.cell(row=1, column=4).font = METADATA_VALUE_FONT

    ws.cell(row=2, column=1, value="Disclosure Scope").font = METADATA_LABEL_FONT
    ws.cell(row=2, column=2).font = METADATA_VALUE_FONT

    # Data validation for disclosure scope
    dv_scope = DataValidation(
        type="list", formula1='"internal,partner,public"', allow_blank=True
    )
    dv_scope.promptTitle = "Disclosure Scope"
    dv_scope.prompt = "Who will see this file? (default: partner)"
    dv_scope.showInputMessage = True
    ws.add_data_validation(dv_scope)
    dv_scope.add("B2")

    # ── Column headers (row 4) ──────────────────────────────────────────────

    headers = [
        "supplier_name",        # A  REQUIRED
        "supplier_id",          # B  optional dedup key
        "jurisdiction",         # C  ISO 3166-1 alpha-2
        "tier",                 # D  1, 2, or 3 (default: 1)
        "parent_supplier",      # E  name or supplier_id of tier N-1 supplier (tier 2/3)
        "business_unit",        # F  optional BU context
        "commodity",            # G  what they supply
        "valid_from",           # H  relationship start (YYYY-MM-DD)
        "annual_value",         # I
        "value_currency",       # J  ISO 4217
        "contract_ref",         # K
        "lei",                  # L
        "duns",                 # M
        "vat",                  # N
        "vat_country",          # O  ISO 3166-1 alpha-2
        "internal_id",          # P
        "risk_tier",            # Q  label (stored on edge)
        "kraljic_quadrant",     # R  label (stored on edge)
        "approval_status",      # S  label (stored on edge)
        "notes",                # T  free text (not imported into graph)
    ]

    HEADER_ROW = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = SUPPLIER_LIST_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[HEADER_ROW].height = 30
    ws.auto_filter.ref = f"A{HEADER_ROW}:T{HEADER_ROW}"

    # ── Data validations ────────────────────────────────────────────────────

    dv_tier = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    dv_tier.promptTitle = "Tier"
    dv_tier.prompt = "Supply-chain tier: 1 = direct, 2 = sub-supplier, 3 = sub-sub-supplier"
    dv_tier.showInputMessage = True
    dv_tier.showErrorMessage = True
    ws.add_data_validation(dv_tier)
    dv_tier.add("D5:D10000")

    dv_risk = DataValidation(
        type="list", formula1='"critical,high,medium,low"', allow_blank=True
    )
    dv_risk.promptTitle = "Risk Tier"
    dv_risk.prompt = "General risk classification"
    dv_risk.showInputMessage = True
    ws.add_data_validation(dv_risk)
    dv_risk.add("Q5:Q10000")

    dv_kraljic = DataValidation(
        type="list",
        formula1='"strategic,leverage,bottleneck,non-critical"',
        allow_blank=True,
    )
    dv_kraljic.promptTitle = "Kraljic Quadrant"
    dv_kraljic.prompt = "Kraljic portfolio classification"
    dv_kraljic.showInputMessage = True
    ws.add_data_validation(dv_kraljic)
    dv_kraljic.add("R5:R10000")

    dv_approval = DataValidation(
        type="list",
        formula1='"approved,conditional,pending,blocked,phase-out"',
        allow_blank=True,
    )
    dv_approval.promptTitle = "Approval Status"
    dv_approval.prompt = "Supplier approval status"
    dv_approval.showInputMessage = True
    ws.add_data_validation(dv_approval)
    dv_approval.add("S5:S10000")

    # ── Column widths ───────────────────────────────────────────────────────

    set_col_widths(ws, {
        "A": 30, "B": 14, "C": 14, "D": 8, "E": 30, "F": 18,
        "G": 20, "H": 14, "I": 14, "J": 14, "K": 16, "L": 24,
        "M": 14, "N": 20, "O": 14, "P": 16, "Q": 12, "R": 18,
        "S": 16, "T": 30,
    })

    # Metadata cell comments
    ws.cell(row=1, column=1).comment = Comment("Your organization's legal name.", "OMTS")
    ws.cell(row=1, column=3).comment = Comment(
        "ISO 8601 date (YYYY-MM-DD) when this supplier list was produced.", "OMTS"
    )
    ws.cell(row=2, column=1).comment = Comment(
        "Who will see this file: internal, partner, or public.", "OMTS"
    )

    # Column header comments
    add_header_comments(ws, {
        1: "Required. Legal name of the supplier organization.",
        2: "Optional dedup key. Rows sharing the same supplier_id collapse to a single organization node, even if names differ. Use when different business units refer to the same supplier by different names.",
        3: "ISO 3166-1 alpha-2 country code of incorporation (e.g. GB, DE, CN).",
        4: "Supply-chain tier: 1 = direct supplier, 2 = sub-supplier, 3 = sub-sub-supplier. Defaults to 1 if blank.",
        5: "For tier 2/3 suppliers: name or supplier_id of the tier N-1 supplier they supply through. Must match a supplier_name or supplier_id in another row.",
        6: "Optional. Your internal business unit that manages this supplier relationship. Allows the same supplier to appear on multiple rows with different risk profiles per BU.",
        7: "HS/CN code or description of what this supplier provides (e.g. 7318.15). Each row represents one supply relationship \u2014 use multiple rows for multiple commodities.",
        8: "ISO 8601 date (YYYY-MM-DD) when this supply relationship started.",
        9: "Annual procurement spend for this relationship (numeric).",
        10: "ISO 4217 currency code for annual_value (e.g. EUR, USD, GBP).",
        11: "Reference number of the governing contract or master service agreement.",
        12: "Legal Entity Identifier (20-character alphanumeric, ISO 17442).",
        13: "D-U-N-S Number (9 digits, Dun & Bradstreet).",
        14: "VAT or tax identification number.",
        15: "ISO 3166-1 alpha-2 country that issued the VAT number.",
        16: "Your internal system identifier for this supplier (e.g. SAP vendor number).",
        17: "Risk classification for this relationship: critical, high, medium, or low. Stored per-relationship (edge), not per-supplier.",
        18: "Kraljic portfolio classification: strategic, leverage, bottleneck, or non-critical. Stored per-relationship (edge).",
        19: "Supplier approval status for this relationship: approved, conditional, pending, blocked, or phase-out.",
        20: "Free-text notes. Not imported into the OMTS graph \u2014 for human reference only.",
    }, header_row=HEADER_ROW)

    return wb


def populate_supplier_list_example(wb):
    """Populate the supplier list with a multi-BU procurement scenario.

    Scenario: Acme Manufacturing's direct and tier-2/3 supplier list for
    steel fastener procurement across Procurement and Engineering BUs.

    Rows 5 and 8 share supplier_id=bolt-001, demonstrating dedup: one org
    node, two supply edges with different business units and risk profiles.
    Row 9 uses parent_supplier=bolt-001 (resolving by supplier_id).
    """
    ws = wb["Supplier List"]

    # Metadata
    ws["B1"] = "Acme Manufacturing GmbH"
    ws["D1"] = "2026-02-22"
    ws["B2"] = "partner"

    # Row 5: Tier 1 — Bolt Supplies via Procurement BU
    ws["A5"] = "Bolt Supplies Ltd"
    ws["B5"] = "bolt-001"
    ws["C5"] = "GB"
    ws["D5"] = 1
    ws["F5"] = "Procurement"
    ws["G5"] = "7318.15"
    ws["H5"] = "2023-01-15"
    ws["I5"] = 450000
    ws["J5"] = "EUR"
    ws["K5"] = "MSA-2023-001"
    ws["M5"] = "234567890"
    ws["Q5"] = "low"
    ws["R5"] = "strategic"
    ws["S5"] = "approved"

    # Row 6: Tier 1 — Nordic Fasteners via Procurement BU
    ws["A6"] = "Nordic Fasteners AB"
    ws["C6"] = "SE"
    ws["D6"] = 1
    ws["F6"] = "Procurement"
    ws["G6"] = "7318.15"
    ws["H6"] = "2024-06-01"
    ws["I6"] = 120000
    ws["J6"] = "EUR"
    ws["Q6"] = "medium"
    ws["R6"] = "leverage"
    ws["S6"] = "conditional"
    ws["T6"] = "Under evaluation; trial order in progress"

    # Row 7: Tier 1 — Shanghai Steel via Procurement BU
    ws["A7"] = "Shanghai Steel Components Co"
    ws["B7"] = "shan-001"
    ws["C7"] = "CN"
    ws["D7"] = 1
    ws["F7"] = "Procurement"
    ws["G7"] = "7228.70"
    ws["H7"] = "2022-03-01"
    ws["I7"] = 800000
    ws["J7"] = "USD"
    ws["K7"] = "FWA-2022-008"
    ws["P7"] = "V-200891"
    ws["Q7"] = "high"
    ws["R7"] = "bottleneck"
    ws["S7"] = "approved"

    # Row 8: Tier 1 — Bolt Supplies again, via Engineering BU (same supplier_id → dedup)
    ws["A8"] = "Bolt Supplies Ltd"
    ws["B8"] = "bolt-001"
    ws["C8"] = "GB"
    ws["D8"] = 1
    ws["F8"] = "Engineering"
    ws["G8"] = "7318.16"
    ws["H8"] = "2024-01-01"
    ws["I8"] = 180000
    ws["J8"] = "EUR"
    ws["Q8"] = "medium"
    ws["R8"] = "non-critical"
    ws["S8"] = "approved"

    # Row 9: Tier 2 — sub-supplier of Bolt (referencing parent by supplier_id)
    ws["A9"] = "Yorkshire Steel Works"
    ws["C9"] = "GB"
    ws["D9"] = 2
    ws["E9"] = "bolt-001"
    ws["G9"] = "7208.10"
    ws["H9"] = "2021-09-01"
    ws["Q9"] = "low"
    ws["S9"] = "approved"

    # Row 10: Tier 2 — sub-supplier of Shanghai Steel (referencing by supplier_id)
    ws["A10"] = "Baosteel Trading Co"
    ws["C10"] = "CN"
    ws["D10"] = 2
    ws["E10"] = "shan-001"
    ws["G10"] = "7207.11"
    ws["H10"] = "2020-01-15"
    ws["Q10"] = "high"
    ws["T10"] = "Primary raw material supplier for Shanghai Steel"

    # Row 11: Tier 3 — sub-supplier of Baosteel (referencing by name, no supplier_id)
    ws["A11"] = "Inner Mongolia Mining Corp"
    ws["C11"] = "CN"
    ws["D11"] = 3
    ws["E11"] = "Baosteel Trading Co"
    ws["G11"] = "2601.11"
    ws["Q11"] = "critical"
    ws["T11"] = "Iron ore source; LKSG high-risk region"


# ── Main ─────────────────────────────────────────────────────────────────────

def create_workbook():
    """Create the OMTS import workbook with all sheets."""
    wb = Workbook()

    create_metadata_sheet(wb)
    create_organizations_sheet(wb)
    create_facilities_sheet(wb)
    create_goods_sheet(wb)
    create_attestations_sheet(wb)
    create_persons_sheet(wb)
    create_consignments_sheet(wb)
    create_supply_relationships_sheet(wb)
    create_corporate_structure_sheet(wb)
    create_same_as_sheet(wb)
    create_identifiers_sheet(wb)
    create_readme_sheet(wb)

    return wb


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Generate empty full template
    wb_template = create_workbook()
    template_path = os.path.join(script_dir, "omts-import-template.xlsx")
    wb_template.save(template_path)
    print(f"Created: {template_path}")

    # Generate full template with example data
    wb_example = create_workbook()
    populate_example_data(wb_example)
    example_path = os.path.join(script_dir, "omts-import-example.xlsx")
    wb_example.save(example_path)
    print(f"Created: {example_path}")

    # Generate empty supplier list template
    wb_sl_template = create_supplier_list_workbook()
    sl_template_path = os.path.join(script_dir, "omts-supplier-list-template.xlsx")
    wb_sl_template.save(sl_template_path)
    print(f"Created: {sl_template_path}")

    # Generate supplier list with example data
    wb_sl_example = create_supplier_list_workbook()
    populate_supplier_list_example(wb_sl_example)
    sl_example_path = os.path.join(script_dir, "omts-supplier-list-example.xlsx")
    wb_sl_example.save(sl_example_path)
    print(f"Created: {sl_example_path}")


if __name__ == "__main__":
    main()
